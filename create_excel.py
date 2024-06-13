from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def create_doc(json_data, file_name):
    try:
        wb = load_workbook(file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Client", 'Anger', 'Sadness', 'Happiness', 'Disgust', 'Fear'])

        ws.column_dimensions[get_column_letter(1)].width = 15
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 10

    colore = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = colore

    for client_id in json_data:
        ws.append([
            client_id, 
            json_data[client_id].get('Anger', 0), 
            json_data[client_id].get('Sadness', 0), 
            json_data[client_id].get('Happiness', 0), 
            json_data[client_id].get('Disgust', 0), 
            json_data[client_id].get('Fear', 0)
        ])

    wb.save(file_name)